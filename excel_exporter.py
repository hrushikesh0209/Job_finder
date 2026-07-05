"""
Excel exporter — produces a formatted, colour-coded workbook with:
  • Jobs sheet  — one row per job, hyperlinked Apply button
  • Summary sheet — stats, top companies, score distribution
"""

import os
import logging
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Palette ──────────────────────────────────────────────────────────────────
NAVY       = "1F4E79"
LIGHT_BLUE = "BDD7EE"
WHITE      = "FFFFFF"
GREEN_BG   = "C6EFCE";  GREEN_FG  = "276221"
YELLOW_BG  = "FFEB9C";  YELLOW_FG = "9C6500"
RED_BG     = "FFC7CE";  RED_FG    = "9C0006"
STRIPE     = "F2F7FF"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _xl_safe(v):
    """Formula-injection guard: openpyxl treats strings starting with '='
    as live formulas, and Excel additionally interprets leading + - @ \\t.
    Job titles/companies/descriptions come from scraped pages, so they are
    attacker-controlled — prefix with ' (Excel renders it as plain text)."""
    if isinstance(v, str) and v[:1] in ("=", "+", "-", "@", "\t"):
        return "'" + v
    return v


def _border() -> Border:
    s = Side(style="thin", color="C8C8C8")
    return Border(left=s, right=s, top=s, bottom=s)


def _font(bold=False, size=10, color="000000", underline=False) -> Font:
    return Font(name="Calibri", bold=bold, size=size,
                color=color, underline="single" if underline else None)


# ──────────────────────────────────────────────────────────────────────────────

def export_to_excel(
    jobs_df: pd.DataFrame,
    match_scores: list = None,
    output_path: str = None,
) -> str:
    """
    Write a formatted Excel file from the jobs DataFrame.

    Args:
        jobs_df:       pandas DataFrame (from jobspy or any source).
        match_scores:  list of dicts with keys:
                         tfidf_score (float 0-100),
                         matched_skills (list[str]),
                         missing_skills (list[str]).
                       Pass None to omit CV columns.
        output_path:   Where to save. Auto-named if None.

    Returns:
        Absolute path to the saved .xlsx file.
    """
    wb = Workbook()
    _build_jobs_sheet(wb.active, jobs_df, match_scores)
    _build_summary_sheet(wb, jobs_df, match_scores)

    if not output_path:
        ts = datetime.now().strftime("%Y-%m-%d_%H%M")
        output_path = f"jobs_{ts}.xlsx"

    parent = os.path.dirname(output_path)
    if parent:
        os.makedirs(parent, exist_ok=True)

    wb.save(output_path)
    logger.info(f"Excel saved → {output_path}")
    return os.path.abspath(output_path)


# ── Jobs sheet ────────────────────────────────────────────────────────────────

def _build_jobs_sheet(ws, df: pd.DataFrame, match_scores):
    ws.title = "Jobs"
    has_scores = match_scores is not None and len(match_scores) == len(df)

    # Column definitions: (header, source_key_or_None, col_width)
    COLS = [
        ("#",                None,                4),
        ("Title",            "title",             30),
        ("Company",          "company",           22),
        ("Location",         "location",          20),
        ("Site",             "site",              12),
        ("Posted",           "date_posted",       13),
        ("Job Type",         "job_type",          13),
        ("Seniority",        "job_level",         14),
        ("Remote",           "is_remote",          8),
        ("Salary",           "__salary__",        18),
    ]
    if has_scores:
        COLS += [
            ("Match %",       "__score__",         10),
            ("Matched Skills","__matched__",        36),
            ("Missing Skills","__missing__",        36),
        ]
    COLS += [
        ("Description",      "description",       55),
        ("Apply",            "__link__",          10),
    ]

    # ── Header ──────────────────────────────────────────────────────────────
    for ci, (hdr, _, w) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci, value=hdr)
        c.fill   = _fill(NAVY)
        c.font   = _font(bold=True, size=11, color=WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 30

    # ── Data rows ────────────────────────────────────────────────────────────
    for ri, (_, job) in enumerate(df.iterrows(), 2):
        sc = match_scores[ri - 2] if has_scores else {}
        pct = sc.get("tfidf_score", 0) if sc else 0

        if has_scores and pct >= 65:
            row_bg = GREEN_BG
        elif has_scores and pct >= 35:
            row_bg = YELLOW_BG
        elif has_scores and pct > 0:
            row_bg = RED_BG
        else:
            row_bg = STRIPE if ri % 2 == 0 else None

        for ci, (hdr, key, _) in enumerate(COLS, 1):
            c = ws.cell(row=ri, column=ci)
            c.border = _border()
            c.font   = _font()
            c.alignment = Alignment(vertical="top", wrap_text=False)

            if row_bg and hdr not in ("Match %",):
                c.fill = _fill(row_bg)

            # ── Cell value logic ─────────────────────────────────────────────
            if hdr == "#":
                c.value = ri - 1
                c.alignment = Alignment(horizontal="center", vertical="top")

            elif hdr == "Salary":
                c.value = _xl_safe(_salary_str(job))

            elif hdr == "Remote":
                c.value = "Yes" if job.get("is_remote") else "No"
                c.alignment = Alignment(horizontal="center", vertical="top")

            elif hdr == "Match %":
                c.value = f"{pct:.0f}%" if pct else "N/A"
                c.alignment = Alignment(horizontal="center", vertical="top")
                if pct >= 65:
                    c.fill = _fill(GREEN_BG); c.font = _font(bold=True, color=GREEN_FG)
                elif pct >= 35:
                    c.fill = _fill(YELLOW_BG); c.font = _font(bold=True, color=YELLOW_FG)
                elif pct > 0:
                    c.fill = _fill(RED_BG); c.font = _font(bold=True, color=RED_FG)

            elif hdr == "Matched Skills":
                c.value = ", ".join((sc or {}).get("matched_skills", []))
                c.alignment = Alignment(vertical="top", wrap_text=True)

            elif hdr == "Missing Skills":
                c.value = ", ".join((sc or {}).get("missing_skills", [])[:10])
                c.alignment = Alignment(vertical="top", wrap_text=True)

            elif hdr == "Description":
                raw = str(job.get("description", "") or "")
                c.value = _xl_safe(raw[:500].replace("\n", " ") + ("…" if len(raw) > 500 else ""))
                c.alignment = Alignment(vertical="top", wrap_text=True)

            elif hdr == "Apply":
                url = str(job.get("job_url") or job.get("job_url_direct") or "")
                # Only http(s) — scraped hrefs could carry javascript:/file: schemes
                if url.lower().startswith(("http://", "https://")):
                    c.value     = "Open ↗"
                    c.hyperlink = url
                    c.font      = _font(color="0563C1", underline=True)
                    c.alignment = Alignment(horizontal="center", vertical="top")

            else:
                # Generic column from df
                c.value = _xl_safe(str(job.get(key, "") or ""))

        ws.row_dimensions[ri].height = 50

    ws.freeze_panes = "B2"


def _salary_str(job) -> str:
    import math as _math
    lo = job.get("min_amount")
    hi = job.get("max_amount")
    def _valid(v):
        return v is not None and not (isinstance(v, float) and _math.isnan(v))
    if not _valid(lo) and not _valid(hi):
        # Boards like Remotive provide salary only as free text
        st = job.get("salary_text")
        return st.strip() if isinstance(st, str) else ""
    curr     = str(job.get("currency") or "")
    interval = str(job.get("interval") or "")
    parts = [str(int(x)) for x in [lo, hi] if _valid(x)]
    return f"{curr}{' – '.join(parts)} {interval}".strip()


# ── Summary sheet ─────────────────────────────────────────────────────────────

def _build_summary_sheet(wb, df: pd.DataFrame, match_scores):
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20

    def title_cell(row, txt):
        c = ws.cell(row=row, column=1, value=txt)
        c.font = _font(bold=True, size=14, color=NAVY)
        ws.merge_cells(f"A{row}:B{row}")
        return c

    def hdr(row):
        for ci, v in enumerate(["Metric", "Value"], 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill   = _fill(LIGHT_BLUE)
            c.font   = _font(bold=True, size=11)
            c.border = _border()
            c.alignment = Alignment(horizontal="center")

    def row_data(row, label, value):
        for ci, v in enumerate([label, value], 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font   = _font(size=10)
            c.border = _border()
            if ci == 2:
                c.alignment = Alignment(horizontal="center")

    # Title
    title_cell(1, "Job Search Summary Report")
    ts = ws.cell(row=2, column=1,
                 value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ts.font = _font(size=9, color="888888")
    ws.merge_cells("A2:B2")

    # Stats table
    hdr(4)
    stats = [
        ("Total Jobs",         len(df)),
        ("Unique Companies",   int(df["company"].nunique()) if "company" in df.columns else "-"),
        ("Sites Searched",     ", ".join(df["site"].unique()) if "site" in df.columns else "-"),
        ("Remote Jobs",        int(df["is_remote"].sum()) if "is_remote" in df.columns else "-"),
    ]

    if match_scores:
        pcts = [s.get("tfidf_score", 0) for s in match_scores if s]
        if pcts:
            avg = sum(pcts) / len(pcts)
            stats += [
                ("Avg CV Match Score", f"{avg:.1f}%"),
                ("High match  ≥65%",   sum(1 for p in pcts if p >= 65)),
                ("Medium match 35–65%",sum(1 for p in pcts if 35 <= p < 65)),
                ("Low match  <35%",    sum(1 for p in pcts if p < 35)),
            ]

    for offset, (lbl, val) in enumerate(stats):
        row_data(5 + offset, lbl, val)

    # Top companies
    if "company" in df.columns:
        start = 5 + len(stats) + 2
        ws.cell(row=start, column=1, value="Top Companies by Listings").font = _font(bold=True, size=12, color=NAVY)
        ws.merge_cells(f"A{start}:B{start}")
        hdr(start + 1)
        for i, (co, cnt) in enumerate(df["company"].value_counts().head(10).items()):
            row_data(start + 2 + i, _xl_safe(str(co)), int(cnt))
